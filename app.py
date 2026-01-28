"""
Fund NAV and Share Roll Calculator - Series Accounting
Form-based input with roll-up logic and detailed output.
"""

import streamlit as st
import pandas as pd
import io

st.set_page_config(page_title="Fund NAV Calculator", page_icon="📊", layout="wide")

st.title("📊 Fund NAV & Share Roll Calculator")
st.markdown("### Series Accounting")
st.markdown("---")

# Initialize session state
if 'num_series' not in st.session_state:
    st.session_state.num_series = 1

# Sidebar configuration
st.sidebar.header("Configuration")

par_value = st.sidebar.number_input(
    "Par Value / New Series Share Price ($)",
    min_value=0.01,
    value=1000.0,
    step=1.0,
    format="%.2f",
    help="Par value for roll-up determination and price for new series shares"
)

# =============================================================================
# STEP 1: Prior Year Ending Balances
# =============================================================================
st.header("Step 1: Prior Year Ending Balances")

col_year1, col_year2, col_spacer = st.columns([1, 1, 4])

with col_year1:
    prior_year = st.number_input(
        "Prior Year",
        min_value=2000,
        max_value=2100,
        value=2023,
        step=1,
        help="The year of the audited financials"
    )

with col_year2:
    current_year = prior_year + 1
    st.text_input(
        "Calculating Year",
        value=str(current_year),
        disabled=True,
        help="Automatically set to prior year + 1"
    )

st.markdown("---")
st.markdown("Enter the prior year ending shares and NAV per share for each series.")
st.markdown("*Note: The first series entered is the 'Initial Series' for roll-up purposes.*")

# Add/Remove series buttons
col_add, col_remove, col_spacer = st.columns([1, 1, 4])

with col_add:
    if st.button("➕ Add Series"):
        st.session_state.num_series += 1
        st.rerun()

with col_remove:
    if st.button("➖ Remove Series") and st.session_state.num_series > 1:
        st.session_state.num_series -= 1
        st.rerun()

# Column headers
col1, col2, col3, col4 = st.columns([2, 1.5, 1.5, 1.5])
with col1:
    st.markdown("**Series Name**")
with col2:
    st.markdown("**Ending Shares**")
with col3:
    st.markdown("**NAV per Share**")
with col4:
    st.markdown("**Total NAV**")

def parse_float(val, default=0.0):
    if val is None or str(val).strip() == '':
        return default
    try:
        cleaned = str(val).replace(',', '').replace('$', '').strip()
        return float(cleaned)
    except:
        return default

prior_series_inputs = []

for i in range(st.session_state.num_series):
    col1, col2, col3, col4 = st.columns([2, 1.5, 1.5, 1.5])

    with col1:
        default_name = "Initial Series" if i == 0 else (f"Series {chr(64 + i)}" if i < 26 else f"Series {i}")
        name = st.text_input(
            "Series Name",
            key=f"prior_series_name_{i}",
            value=default_name,
            label_visibility="collapsed",
            placeholder="Series Name"
        )
        if i == 0:
            st.caption("(Initial Series)")

    with col2:
        shares_str = st.text_input(
            "Ending Shares",
            key=f"prior_series_shares_{i}",
            value="",
            label_visibility="collapsed",
            placeholder="Enter shares"
        )
        shares = parse_float(shares_str, 0.0)

    with col3:
        nav_str = st.text_input(
            "NAV per Share ($)",
            key=f"prior_series_nav_{i}",
            value="",
            label_visibility="collapsed",
            placeholder="Enter NAV/share"
        )
        nav = parse_float(nav_str, 0.0)

    with col4:
        total = shares * nav
        st.markdown(f"**${total:,.2f}**")

    prior_series_inputs.append({
        'Series': name,
        'Ending Shares': shares,
        'NAV per Share': nav,
        'Total NAV': total,
        'is_initial': i == 0
    })

valid_prior_series = [s for s in prior_series_inputs if s['Series'] and s['Ending Shares'] > 0]

if valid_prior_series:
    prior_year_df = pd.DataFrame(valid_prior_series)
    st.markdown("---")
    col1, col2 = st.columns(2)
    with col1:
        st.metric("Total Prior Year NAV", f"${prior_year_df['Total NAV'].sum():,.2f}")
    with col2:
        st.metric("Number of Prior Year Series", len(prior_year_df))

st.markdown("---")

# =============================================================================
# STEP 2: Monthly Activity
# =============================================================================
st.header(f"Step 2: Monthly Activity for {current_year}")

st.markdown("""
Enter the P/L, contributions, and redemptions for each month.
- **P/L**: Used to calculate NAV per share (not shown in final output)
- **Contributions**: Creates a new series (Series M/YYYY)
- **Redemptions**: Enter amount and select which series
""")

months = ['January', 'February', 'March', 'April', 'May', 'June',
          'July', 'August', 'September', 'October', 'November', 'December']

def get_available_series_up_to_month(month_idx, prior_series, monthly_data, current_year):
    available = [s['Series'] for s in prior_series if s['Ending Shares'] > 0]
    for i in range(month_idx):
        contrib = monthly_data[i]['contributions']
        if contrib > 0:
            month_num = i + 1
            new_series_name = f"Series {month_num}/{current_year}"
            available.append(new_series_name)
    return available

monthly_data = []

st.markdown("---")
col_month, col_pl, col_contrib, col_redemp, col_full, col_series = st.columns([1.2, 1.2, 1.2, 1.2, 0.8, 1.8])
with col_month:
    st.markdown("**Month**")
with col_pl:
    st.markdown("**P/L**")
with col_contrib:
    st.markdown("**Contrib**")
with col_redemp:
    st.markdown("**Redemp $**")
with col_full:
    st.markdown("**Full?**")
with col_series:
    st.markdown("**From Series**")

for i, month in enumerate(months):
    col_month, col_pl, col_contrib, col_redemp, col_full, col_series = st.columns([1.2, 1.2, 1.2, 1.2, 0.8, 1.8])

    with col_month:
        st.markdown(f"**{month}**")

    with col_pl:
        pl_str = st.text_input("P/L", key=f"pl_{i}", value="", label_visibility="collapsed", placeholder="0")
        pl = parse_float(pl_str, 0.0)

    with col_contrib:
        contrib_str = st.text_input("Contributions", key=f"contrib_{i}", value="", label_visibility="collapsed", placeholder="0")
        contrib = parse_float(contrib_str, 0.0)

    with col_redemp:
        redemp_str = st.text_input("Redemptions", key=f"redemp_{i}", value="", label_visibility="collapsed", placeholder="0")
        redemp = parse_float(redemp_str, 0.0)

    with col_full:
        full_redemption = st.checkbox("Full", key=f"full_redemp_{i}", help="Check for full redemption of selected series")

    with col_series:
        available_series = get_available_series_up_to_month(i, prior_series_inputs, monthly_data, current_year)
        if (redemp > 0 or full_redemption) and available_series:
            selected_series = st.selectbox("Series", options=available_series, key=f"redemp_series_{i}", label_visibility="collapsed")
        elif redemp > 0 or full_redemption:
            st.markdown("*No series*")
            selected_series = None
        else:
            st.markdown("—")
            selected_series = None

    monthly_data.append({
        'month': month,
        'month_num': i + 1,
        'pl': pl,
        'contributions': contrib,
        'redemptions': redemp,
        'redemption_series': selected_series,
        'full_redemption': full_redemption
    })

st.markdown("---")

# =============================================================================
# STEP 3: Calculate
# =============================================================================
st.header("Step 3: Calculate Share Roll & NAV")

if st.button("🔄 Calculate Share Roll", type="primary", use_container_width=True):

    if not valid_prior_series:
        st.error("Please enter at least one prior year series with shares > 0")
    else:
        # Initialize series tracking with beginning of year values
        series_data = {}
        initial_series_name = valid_prior_series[0]['Series']

        # Track all series that ever existed (for output)
        all_series_ever = set()

        for s in valid_prior_series:
            series_data[s['Series']] = {
                'beginning_shares': s['Ending Shares'],
                'beginning_nav': s['NAV per Share'],
                'shares': s['Ending Shares'],
                'nav_per_share': s['NAV per Share'],
                'total_nav': s['Total NAV'],
                'transfers_in': 0.0,
                'transfers_out': 0.0,
                'contributed_shares': 0.0,
                'redeemed_shares': 0.0,
                'is_initial': s.get('is_initial', False),
                'created_month': None,
                'rolled_up': False
            }
            all_series_ever.add(s['Series'])

        # Detailed calculation log
        calc_log = []

        # =====================================================================
        # ROLL-UP LOGIC: At beginning of year
        # =====================================================================
        calc_log.append({
            'Step': 'Roll-up Check',
            'Month': 'Beginning of Year',
            'Series': 'All',
            'Description': f'Checking if any series NAV > par value (${par_value:,.2f})',
            'Details': ''
        })

        # Find series that need to roll up (NAV > par value, not the initial series)
        rollup_series = []
        for series_name, s in series_data.items():
            if not s['is_initial'] and s['nav_per_share'] > par_value and s['shares'] > 0:
                rollup_series.append(series_name)

        initial_series = series_data.get(initial_series_name)

        if rollup_series and initial_series and initial_series['shares'] > 0:
            for series_name in rollup_series:
                s = series_data[series_name]

                # Calculate transfer
                transfer_value = s['shares'] * s['nav_per_share']
                shares_transferred_out = s['shares']
                shares_transferred_in = transfer_value / initial_series['nav_per_share'] if initial_series['nav_per_share'] > 0 else 0

                calc_log.append({
                    'Step': 'Roll-up Transfer',
                    'Month': 'Beginning of Year',
                    'Series': series_name,
                    'Description': f'Rolling up into {initial_series_name}',
                    'Details': f'Shares out: {shares_transferred_out:,.4f} @ ${s["nav_per_share"]:,.4f} = ${transfer_value:,.2f}'
                })

                calc_log.append({
                    'Step': 'Roll-up Transfer',
                    'Month': 'Beginning of Year',
                    'Series': initial_series_name,
                    'Description': f'Receiving roll-up from {series_name}',
                    'Details': f'Shares in: ${transfer_value:,.2f} / ${initial_series["nav_per_share"]:,.4f} = {shares_transferred_in:,.4f} shares'
                })

                # Update series data
                s['transfers_out'] = shares_transferred_out
                s['shares'] = 0
                s['total_nav'] = 0
                s['rolled_up'] = True

                initial_series['transfers_in'] += shares_transferred_in
                initial_series['shares'] += shares_transferred_in
                initial_series['total_nav'] += transfer_value
        else:
            calc_log.append({
                'Step': 'Roll-up Check',
                'Month': 'Beginning of Year',
                'Series': 'All',
                'Description': 'No roll-ups required',
                'Details': 'No series with NAV > par value'
            })

        # =====================================================================
        # MONTHLY PROCESSING
        # =====================================================================
        for month_info in monthly_data:
            month = month_info['month']
            month_num = month_info['month_num']
            pl = month_info['pl']
            contributions = month_info['contributions']
            redemptions = month_info['redemptions']
            redemption_series = month_info['redemption_series']

            # 1. Create new series from contributions
            full_redemption = month_info['full_redemption']

            if contributions > 0:
                new_series_name = f"Series {month_num}/{current_year}"
                counter = 1
                base_name = new_series_name
                while new_series_name in series_data:
                    counter += 1
                    new_series_name = f"{base_name}-{counter}"

                new_shares = contributions / par_value

                calc_log.append({
                    'Step': 'New Series',
                    'Month': month,
                    'Series': new_series_name,
                    'Description': f'Contribution of ${contributions:,.2f}',
                    'Details': f'Shares issued: ${contributions:,.2f} / ${par_value:,.2f} = {new_shares:,.4f}'
                })

                series_data[new_series_name] = {
                    'beginning_shares': 0.0,
                    'beginning_nav': par_value,
                    'shares': new_shares,
                    'nav_per_share': par_value,
                    'total_nav': contributions,
                    'transfers_in': 0.0,
                    'transfers_out': 0.0,
                    'contributed_shares': new_shares,
                    'redeemed_shares': 0.0,
                    'is_initial': False,
                    'created_month': month,
                    'rolled_up': False
                }
                all_series_ever.add(new_series_name)

            # 2. Allocate P/L pro-rata (AFTER contributions, BEFORE redemptions)
            # Build explicit list of active series to avoid any dict iteration issues
            active_series_for_pl = [(name, data) for name, data in series_data.items()
                                    if data['shares'] > 0 and data['total_nav'] > 0]
            total_nav_for_pl = sum(data['total_nav'] for _, data in active_series_for_pl)

            if total_nav_for_pl > 0 and pl != 0:
                # Snapshot NAV values before any modifications
                nav_snapshot = {name: data['total_nav'] for name, data in active_series_for_pl}

                calc_log.append({
                    'Step': 'P/L Allocation',
                    'Month': month,
                    'Series': 'All',
                    'Description': f'Total P/L: ${pl:,.2f}',
                    'Details': f'Total NAV for allocation: ${total_nav_for_pl:,.2f}'
                })

                for series_name, _ in active_series_for_pl:
                    s = series_data[series_name]
                    pl_share = pl * (nav_snapshot[series_name] / total_nav_for_pl)
                    old_nav = s['nav_per_share']
                    s['total_nav'] += pl_share
                    if s['shares'] > 0:
                        s['nav_per_share'] = s['total_nav'] / s['shares']

                    calc_log.append({
                        'Step': 'P/L Allocation',
                        'Month': month,
                        'Series': series_name,
                        'Description': f'P/L share: ${pl_share:,.2f}',
                        'Details': f'NAV/share: ${old_nav:,.4f} → ${s["nav_per_share"]:,.4f}'
                    })

            # 3. Process redemption (AFTER P/L - at post-P/L NAV)
            if (redemptions > 0 or full_redemption) and redemption_series and redemption_series in series_data:
                s = series_data[redemption_series]
                if s['nav_per_share'] > 0 and s['shares'] > 0:
                    if full_redemption:
                        shares_redeemed = s['shares']
                        redemption_amount = shares_redeemed * s['nav_per_share']

                        calc_log.append({
                            'Step': 'Full Redemption',
                            'Month': month,
                            'Series': redemption_series,
                            'Description': f'FULL redemption of all shares',
                            'Details': f'NAV/share: ${s["nav_per_share"]:,.4f} | Shares redeemed: {shares_redeemed:,.4f} | Redemption value: ${redemption_amount:,.2f}'
                        })
                    else:
                        shares_redeemed = redemptions / s['nav_per_share']
                        redemption_amount = redemptions

                        calc_log.append({
                            'Step': 'Redemption',
                            'Month': month,
                            'Series': redemption_series,
                            'Description': f'Redemption of ${redemptions:,.2f}',
                            'Details': f'NAV/share: ${s["nav_per_share"]:,.4f} | Shares redeemed: ${redemptions:,.2f} / ${s["nav_per_share"]:,.4f} = {shares_redeemed:,.4f} | Shares before: {s["shares"]:,.4f}'
                        })

                    s['redeemed_shares'] += shares_redeemed
                    s['shares'] = s['shares'] - shares_redeemed
                    s['total_nav'] = s['total_nav'] - redemption_amount

        # =====================================================================
        # BUILD OUTPUT
        # =====================================================================
        output_rows = []

        for series_name in all_series_ever:
            s = series_data[series_name]
            output_rows.append({
                'Series': series_name,
                'Beginning Shares': s['beginning_shares'],
                'Transfers In': s['transfers_in'],
                'Transfers Out': s['transfers_out'],
                'Contributed Shares': s['contributed_shares'],
                'Redeemed Shares': s['redeemed_shares'],
                'Ending Shares': s['shares'],
                'Ending NAV per Share': s['nav_per_share'] if s['shares'] > 0 else 0.0
            })

        # Sort: Initial series first, then prior year series, then new series by month
        def sort_key(row):
            name = row['Series']
            if name == initial_series_name:
                return (0, name)
            elif '/' not in name:  # Prior year series
                return (1, name)
            else:  # New series
                return (2, name)

        output_rows.sort(key=sort_key)

        # Add total row
        total_row = {
            'Series': 'TOTAL',
            'Beginning Shares': sum(r['Beginning Shares'] for r in output_rows),
            'Transfers In': sum(r['Transfers In'] for r in output_rows),
            'Transfers Out': sum(r['Transfers Out'] for r in output_rows),
            'Contributed Shares': sum(r['Contributed Shares'] for r in output_rows),
            'Redeemed Shares': sum(r['Redeemed Shares'] for r in output_rows),
            'Ending Shares': sum(r['Ending Shares'] for r in output_rows),
            'Ending NAV per Share': ''  # N/A for total
        }
        output_rows.append(total_row)

        output_df = pd.DataFrame(output_rows)
        calc_log_df = pd.DataFrame(calc_log)

        # =====================================================================
        # DISPLAY RESULTS
        # =====================================================================
        st.markdown("---")
        st.header("📈 Share Roll Summary")

        # Format for display
        display_df = output_df.copy()

        for col in ['Beginning Shares', 'Transfers In', 'Transfers Out', 'Contributed Shares', 'Redeemed Shares', 'Ending Shares']:
            display_df[col] = display_df[col].apply(lambda x: f"{x:,.4f}" if isinstance(x, (int, float)) else x)

        display_df['Ending NAV per Share'] = display_df['Ending NAV per Share'].apply(
            lambda x: f"${x:,.4f}" if isinstance(x, (int, float)) and x > 0 else "—" if x == '' else "—"
        )

        # Style the dataframe
        st.dataframe(display_df, use_container_width=True, hide_index=True)

        # Summary metrics
        st.markdown("---")
        col1, col2, col3, col4 = st.columns(4)

        total_beginning = sum(r['Beginning Shares'] for r in output_rows[:-1])
        total_ending = sum(r['Ending Shares'] for r in output_rows[:-1])
        total_contributions = sum(r['Contributed Shares'] for r in output_rows[:-1])
        total_redemptions = sum(r['Redeemed Shares'] for r in output_rows[:-1])
        total_transfers_in = sum(r['Transfers In'] for r in output_rows[:-1])
        total_transfers_out = sum(r['Transfers Out'] for r in output_rows[:-1])

        with col1:
            st.metric("Beginning Shares", f"{total_beginning:,.4f}")
        with col2:
            st.metric("Contributed Shares", f"{total_contributions:,.4f}")
        with col3:
            st.metric("Redeemed Shares", f"{total_redemptions:,.4f}")
        with col4:
            st.metric("Ending Shares", f"{total_ending:,.4f}")

        # Check figures
        st.markdown("---")
        total_ending_nav = sum(s['total_nav'] for s in series_data.values() if s['shares'] > 0)
        total_year_pl = sum(md['pl'] for md in monthly_data)
        col_check1, col_check2 = st.columns(2)
        with col_check1:
            st.metric("Total Ending NAV (Check Figure)", f"${total_ending_nav:,.2f}")
        with col_check2:
            st.metric("Total P/L for Year (Check Figure)", f"${total_year_pl:,.2f}")

        # Reconciliation check
        calculated_ending = total_beginning + total_transfers_in - total_transfers_out + total_contributions - total_redemptions
        if abs(calculated_ending - total_ending) > 0.0001:
            st.warning(f"⚠️ Reconciliation difference: {calculated_ending:,.4f} calculated vs {total_ending:,.4f} reported (diff: {calculated_ending - total_ending:,.4f})")
        else:
            st.success("✓ Share roll reconciles: Beginning + Transfers In - Transfers Out + Contributed - Redeemed = Ending")

        # Monthly NAV per Share tracking
        st.subheader("Monthly NAV per Share by Series")
        st.markdown("*Use these values to verify redemption amounts*")

        # Build monthly NAV tracking
        monthly_nav_data = []

        # Re-run calculation just to capture monthly NAV snapshots
        temp_series = {}
        for s in valid_prior_series:
            temp_series[s['Series']] = {
                'shares': s['Ending Shares'],
                'nav_per_share': s['NAV per Share'],
                'total_nav': s['Total NAV'],
            }

        # Process roll-ups first
        temp_initial = temp_series.get(initial_series_name)
        for series_name, s in list(temp_series.items()):
            if series_name != initial_series_name and s['nav_per_share'] > par_value and s['shares'] > 0:
                transfer_value = s['shares'] * s['nav_per_share']
                shares_in = transfer_value / temp_initial['nav_per_share'] if temp_initial['nav_per_share'] > 0 else 0
                temp_initial['shares'] += shares_in
                temp_initial['total_nav'] += transfer_value
                s['shares'] = 0
                s['total_nav'] = 0

        # Capture beginning of year NAV
        boy_row = {'Month': 'Beginning of Year'}
        for series_name, s in temp_series.items():
            if s['shares'] > 0:
                boy_row[series_name] = s['nav_per_share']
        monthly_nav_data.append(boy_row)

        # Process each month
        for month_info in monthly_data:
            month = month_info['month']
            pl = month_info['pl']
            contributions = month_info['contributions']
            redemptions = month_info['redemptions']
            redemption_series = month_info['redemption_series']

            # 1. Create new series (contributions first)
            full_redemption = month_info['full_redemption']
            if contributions > 0:
                month_num = month_info['month_num']
                new_name = f"Series {month_num}/{current_year}"
                temp_series[new_name] = {
                    'shares': contributions / par_value,
                    'nav_per_share': par_value,
                    'total_nav': contributions
                }

            # 2. Allocate P/L (after contributions, before redemptions)
            # Build explicit list of active series
            active_temp = [(name, data) for name, data in temp_series.items()
                          if data['shares'] > 0 and data['total_nav'] > 0]
            total_nav = sum(data['total_nav'] for _, data in active_temp)
            if total_nav > 0 and pl != 0:
                snapshot = {name: data['total_nav'] for name, data in active_temp}
                for series_name, _ in active_temp:
                    s = temp_series[series_name]
                    if snapshot[series_name] > 0 and s['shares'] > 0:
                        pl_share = pl * (snapshot[series_name] / total_nav)
                        s['total_nav'] += pl_share
                        s['nav_per_share'] = s['total_nav'] / s['shares']

            # 3. Process redemption (after P/L)
            if (redemptions > 0 or full_redemption) and redemption_series and redemption_series in temp_series:
                s = temp_series[redemption_series]
                if s['nav_per_share'] > 0 and s['shares'] > 0:
                    if full_redemption:
                        redemption_amount = s['shares'] * s['nav_per_share']
                        s['total_nav'] -= redemption_amount
                        s['shares'] = 0
                    else:
                        shares_red = redemptions / s['nav_per_share']
                        s['shares'] -= shares_red
                        s['total_nav'] -= redemptions

            # Capture end of month NAV
            month_row = {'Month': f'End of {month}'}
            for series_name, s in temp_series.items():
                if s['shares'] > 0:
                    month_row[series_name] = s['nav_per_share']
            monthly_nav_data.append(month_row)

        nav_tracking_df = pd.DataFrame(monthly_nav_data)

        # Format for display
        display_nav_df = nav_tracking_df.copy()
        for col in display_nav_df.columns:
            if col != 'Month':
                display_nav_df[col] = display_nav_df[col].apply(
                    lambda x: f"${x:,.4f}" if pd.notna(x) and isinstance(x, (int, float)) else "—"
                )

        st.dataframe(display_nav_df, use_container_width=True, hide_index=True)

        # Calculation details
        st.subheader("Calculation Details")
        with st.expander("View Step-by-Step Calculations", expanded=False):
            st.dataframe(calc_log_df, use_container_width=True, hide_index=True)

        # =====================================================================
        # EXCEL EXPORT
        # =====================================================================
        st.markdown("---")
        st.subheader("📥 Download Results")

        output_buffer = io.BytesIO()
        with pd.ExcelWriter(output_buffer, engine='openpyxl') as writer:
            # Sheet 1: Summary Output (numbers, not strings)
            # Keep output_df as numbers for Excel
            excel_output_df = output_df.copy()
            # Replace empty string with None for proper Excel handling
            excel_output_df['Ending NAV per Share'] = excel_output_df['Ending NAV per Share'].replace('', None)
            excel_output_df.to_excel(writer, index=False, sheet_name='Share Roll Summary')

            # Format the Excel sheet
            workbook = writer.book
            worksheet = writer.sheets['Share Roll Summary']

            # Apply number format with commas to numeric columns
            from openpyxl.styles import numbers
            for row in range(2, len(excel_output_df) + 2):  # Start from row 2 (after header)
                for col in range(2, 8):  # Columns B through G (shares columns)
                    cell = worksheet.cell(row=row, column=col)
                    if cell.value is not None and isinstance(cell.value, (int, float)):
                        cell.number_format = '#,##0.0000'
                # NAV per Share column (column H)
                cell = worksheet.cell(row=row, column=8)
                if cell.value is not None and isinstance(cell.value, (int, float)):
                    cell.number_format = '$#,##0.0000'

            # Auto-fit column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                worksheet.column_dimensions[column_letter].width = max(max_length + 2, 12)

            # Sheet 2: Calculation Details with Excel Formulas
            # Build a detailed calculation sheet showing month-by-month progression
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter

            calc_sheet = workbook.create_sheet('Calculation Details')

            # Styles
            header_font = Font(bold=True)
            header_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
            month_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
            currency_format = '$#,##0.00'
            shares_format = '#,##0.0000'
            pct_format = '0.00%'
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            row = 1

            # Section 1: Prior Year Series Data
            calc_sheet.cell(row=row, column=1, value='PRIOR YEAR ENDING BALANCES').font = Font(bold=True, size=12)
            row += 1

            headers = ['Series', 'Ending Shares', 'NAV per Share', 'Total NAV']
            for col, header in enumerate(headers, 1):
                cell = calc_sheet.cell(row=row, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
            row += 1

            prior_start_row = row
            for s in valid_prior_series:
                calc_sheet.cell(row=row, column=1, value=s['Series'])
                calc_sheet.cell(row=row, column=2, value=s['Ending Shares']).number_format = shares_format
                calc_sheet.cell(row=row, column=3, value=s['NAV per Share']).number_format = currency_format
                # Formula: Shares * NAV per Share
                calc_sheet.cell(row=row, column=4, value=f'=B{row}*C{row}').number_format = currency_format
                row += 1
            prior_end_row = row - 1

            # Total row
            calc_sheet.cell(row=row, column=1, value='TOTAL').font = header_font
            calc_sheet.cell(row=row, column=2, value=f'=SUM(B{prior_start_row}:B{prior_end_row})').number_format = shares_format
            calc_sheet.cell(row=row, column=4, value=f'=SUM(D{prior_start_row}:D{prior_end_row})').number_format = currency_format
            prior_total_row = row
            row += 2

            # Section 2: Monthly Calculations
            calc_sheet.cell(row=row, column=1, value='MONTHLY CALCULATIONS').font = Font(bold=True, size=12)
            row += 1

            # Track series NAV for formula references
            series_nav_refs = {}  # Will store the cell reference for each series' current NAV

            # Initialize with prior year data
            for idx, s in enumerate(valid_prior_series):
                series_nav_refs[s['Series']] = f'D{prior_start_row + idx}'

            months = ['January', 'February', 'March', 'April', 'May', 'June',
                      'July', 'August', 'September', 'October', 'November', 'December']

            for month_idx, month_info in enumerate(monthly_data):
                month = month_info['month']
                pl = month_info['pl']
                contributions = month_info['contributions']
                redemptions = month_info['redemptions']
                redemption_series = month_info['redemption_series']
                full_redemption = month_info['full_redemption']

                # Skip months with no activity
                if pl == 0 and contributions == 0 and redemptions == 0 and not full_redemption:
                    continue

                # Month header
                calc_sheet.cell(row=row, column=1, value=f'{month}').font = Font(bold=True, size=11)
                calc_sheet.cell(row=row, column=1).fill = month_fill
                for col in range(2, 8):
                    calc_sheet.cell(row=row, column=col).fill = month_fill
                row += 1

                # Column headers: Contributions → P/L → Redemptions (matches calculation order)
                month_headers = ['Series', 'Beginning NAV', 'Contribution', 'NAV pre-P/L', 'P/L %', 'P/L Allocated', 'NAV post-P/L', 'Redemption', 'Ending NAV']
                for col, header in enumerate(month_headers, 1):
                    cell = calc_sheet.cell(row=row, column=col, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                row += 1

                month_start_row = row
                active_series = list(series_nav_refs.keys())

                new_series_refs = {}

                for series_name in active_series:
                    calc_sheet.cell(row=row, column=1, value=series_name)

                    # Beginning NAV
                    beg_nav_ref = series_nav_refs[series_name]
                    calc_sheet.cell(row=row, column=2, value=f'={beg_nav_ref}').number_format = currency_format

                    # Contribution (existing series don't get contributions)
                    calc_sheet.cell(row=row, column=3, value=0).number_format = currency_format

                    # NAV pre-P/L = Beginning + Contribution
                    calc_sheet.cell(row=row, column=4, value=f'=B{row}+C{row}').number_format = currency_format

                    new_series_refs[series_name] = row
                    row += 1

                # Add new series row from contribution
                if contributions > 0:
                    month_num = month_info['month_num']
                    new_series_name = f"Series {month_num}/{current_year}"

                    calc_sheet.cell(row=row, column=1, value=new_series_name)
                    calc_sheet.cell(row=row, column=2, value=0).number_format = currency_format
                    calc_sheet.cell(row=row, column=3, value=contributions).number_format = currency_format
                    calc_sheet.cell(row=row, column=4, value=f'=B{row}+C{row}').number_format = currency_format

                    new_series_refs[new_series_name] = row
                    row += 1

                month_end_row = row - 1

                # Build total NAV pre-P/L formula
                total_nav_pre_pl = f'SUM(D{month_start_row}:D{month_end_row})'

                # Fill in P/L %, P/L Allocated, NAV post-P/L, Redemption, Ending NAV
                for sname, srow in new_series_refs.items():
                    if pl != 0:
                        # P/L % = NAV pre-P/L / Total NAV pre-P/L
                        calc_sheet.cell(row=srow, column=5, value=f'=IF({total_nav_pre_pl}=0,0,D{srow}/{total_nav_pre_pl})').number_format = pct_format
                        # P/L Allocated = P/L % * Total P/L
                        calc_sheet.cell(row=srow, column=6, value=f'=E{srow}*{pl}').number_format = currency_format
                    else:
                        calc_sheet.cell(row=srow, column=5, value=0).number_format = pct_format
                        calc_sheet.cell(row=srow, column=6, value=0).number_format = currency_format

                    # NAV post-P/L = NAV pre-P/L + P/L Allocated
                    calc_sheet.cell(row=srow, column=7, value=f'=D{srow}+F{srow}').number_format = currency_format

                    # Redemption
                    if (redemptions > 0 or full_redemption) and redemption_series == sname:
                        if full_redemption:
                            calc_sheet.cell(row=srow, column=8, value=f'=-G{srow}').number_format = currency_format
                        else:
                            calc_sheet.cell(row=srow, column=8, value=-redemptions).number_format = currency_format
                    else:
                        calc_sheet.cell(row=srow, column=8, value=0).number_format = currency_format

                    # Ending NAV = NAV post-P/L + Redemption
                    calc_sheet.cell(row=srow, column=9, value=f'=G{srow}+H{srow}').number_format = currency_format

                # Update series_nav_refs to point to Ending NAV column (I)
                final_refs = {}
                for sname, srow in new_series_refs.items():
                    final_refs[sname] = f'I{srow}'

                # Total row for month
                calc_sheet.cell(row=row, column=1, value='Month Total').font = header_font
                calc_sheet.cell(row=row, column=2, value=f'=SUM(B{month_start_row}:B{month_end_row})').number_format = currency_format
                calc_sheet.cell(row=row, column=3, value=f'=SUM(C{month_start_row}:C{month_end_row})').number_format = currency_format
                calc_sheet.cell(row=row, column=4, value=f'=SUM(D{month_start_row}:D{month_end_row})').number_format = currency_format
                calc_sheet.cell(row=row, column=6, value=f'=SUM(F{month_start_row}:F{month_end_row})').number_format = currency_format
                calc_sheet.cell(row=row, column=7, value=f'=SUM(G{month_start_row}:G{month_end_row})').number_format = currency_format
                calc_sheet.cell(row=row, column=8, value=f'=SUM(H{month_start_row}:H{month_end_row})').number_format = currency_format
                calc_sheet.cell(row=row, column=9, value=f'=SUM(I{month_start_row}:I{month_end_row})').number_format = currency_format

                # Update series references for next month
                series_nav_refs = final_refs
                row += 2

            # Auto-fit columns
            for col in range(1, 10):
                calc_sheet.column_dimensions[get_column_letter(col)].width = 16

            # Sheet 3: Inputs Summary
            inputs_df = pd.DataFrame({
                'Parameter': ['Prior Year', 'Calculating Year', 'Par Value'],
                'Value': [prior_year, current_year, par_value]
            })

            prior_inputs = pd.DataFrame(valid_prior_series)
            monthly_inputs = pd.DataFrame(monthly_data)

            inputs_df.to_excel(writer, index=False, sheet_name='Inputs', startrow=0)
            prior_inputs.to_excel(writer, index=False, sheet_name='Inputs', startrow=5)
            monthly_inputs.to_excel(writer, index=False, sheet_name='Inputs', startrow=5 + len(prior_inputs) + 3)

        output_buffer.seek(0)

        st.download_button(
            label="📥 Download Excel (Summary + Calculations)",
            data=output_buffer,
            file_name=f"share_roll_{current_year}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

# Footer
st.markdown("---")
st.markdown("""
<div style='text-align: center; color: gray;'>
    <small>Fund NAV & Share Roll Calculator | Series Accounting</small>
</div>
""", unsafe_allow_html=True)
